#google ドライブとつなぐためのライブラリ
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
#インターネットに公開するライブラリ
import streamlit as st
#画像を簡単に処理するライブラリ
from PIL import Image, ImageFilter
#数的処理するライブラリ
import numpy as np
import pickle
import os
import pandas as pd
import csv
import openpyxl

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

#事前準備ss
#ID次の番号取得
folder_id = '10Ogv7m81vckhXxmRdleo5xouy6lO6O7V' 
file_id = drive.ListFile({'q': 'title = "df.csv"'}).GetList()[0]['id']
f = drive.CreateFile({'id': file_id})
f.GetContentFile('df.csv')

#新規時かぶらないように
shoki = sum(1 for line in open('df.csv', 'r')) + 7200 

#body
st.title('栽培データ入力フォーム')
genre = st.radio(
     "何を行いますか？",
     ('新規入力', '編集', 'データEXCEL化'))

#データ取得
with open('df.csv','r') as f:
    #df = pd.read_csv(f, header=0)
    csv_files = csv.reader(f)
    i = 0
    for csv_file in csv_files:
        #st.text(csv.file)
        if i == 0:
            colmuns = csv_file
            i = 1
            #st.text(colmuns)
        elif i==1:
            mylist = [csv_file]
            #st.text(mylist)
            i = 2
            #st.text(csv_file)
        elif i >1 and csv_file is not None:
            #st.text(mylist)
            mylist.append(csv_file)
        else :
            pass

    df = pd.DataFrame(mylist,columns= colmuns)

#新規入力画面
if genre == '新規入力':
    col1, col2 = st.columns(2)
    with col1:
        main_id = st.text_input('id',value=shoki)
    with col2:
        title = st.text_input('タイトル')

    st.markdown('【サンプル１】')
    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        a1 = st.text_input('草丈',key=1)#value＝2　デファルト数入れられる
    with col2:
        b1 = st.text_input('',key=2)
    with col3:    
        c1 = st.text_input('',key=3)
    with col4:
        d1 = st.text_input('',key=4)
    with col5:
        e1 = st.text_input('',key=5)

    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        f1 = st.text_input('茎数',key=6)#value＝2　デファルト数入れられる
    with col2:
        g1 = st.text_input('',key=7)
    with col3:    
        h1 = st.text_input('',key=8)
    with col4:
        i1 = st.text_input('',key=9)
    with col5:
        j1 = st.text_input('',key=10)

    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        k1 = st.text_input('SPAD',key=11)#value＝2　デファルト数入れられる
    with col2:
        l1 = st.text_input('',key=12)
    with col3:    
        m1 = st.text_input('',key=13)
    with col4:
        n1 = st.text_input('',key=14)
    with col5:
        o1 = st.text_input('',key=15)

    field = st.file_uploader('全体写真')
    if field:
        st.image(field)
    close = st.file_uploader('近距離写真')
    if close:
        st.image(close)


    download_name_a = 'field' + main_id
    download_name_b = 'close' + main_id

    button_upload = st.button('データ保存')
    if button_upload:
        #st.markdown(f'{field.name}をアップロードしました。')
        with open(field.name,'wb') as f:
            f.write(field.read()) 
        f = drive.CreateFile({'title':download_name_a,#field.name
                            'mimeType':'image/png,image/jpeg',
                            'parents':[{'id':folder_id}]})
        f.SetContentFile(field.name)
        f.Upload()

        file_id_a = drive.ListFile().GetList()
        fx = file_id_a['title' == download_name_a]['id']
        f.clear()

        #st.markdown(f'{close.name}をアップロードしました。')
        with open(close.name,'wb') as f:
            f.write(close.read())
        f = drive.CreateFile({'title':download_name_b,
                            'mimeType':'image/png,imag/jpeg',
                            'parents':[{'id':folder_id}]})
        f.SetContentFile(close.name)
        f.Upload()
        file_id_b = drive.ListFile().GetList()
        fb = file_id_b['title' == download_name_b]['id']
        f.clear()
        
        #df1 = pd.DataFrame(data = data,columns=colmuns)
        colmuns = ['id','title',
        'kusa1','kusa2','kusa3','kusa4','kusa5',
        'kuki1','kuki2','kuki3','kuki4','kuki5',
        'spad1','spad2','spad3','spad4','spad5',
        'field','close']

        data = [main_id,title,
        a1,b1,c1,d1,e1,
        f1,g1,h1,i1,j1,
        k1,l1,m1,n1,o1,
        fx,fb]
        #st.text(data)
        #ほんとの最初にデータを作る時
#        with open('df.csv','w') as f:
#            writer = csv.writer(f)
#            writer.writerow(colmuns)
#            writer.writerow(data)
#
#        f = drive.CreateFile({'title':'df.csv',
#                            'mimeType':'text/csv',
#                            'parents':[{'id':folder_id}]})
#        f.SetContentFile('df.csv')
#        f.Upload()
#        f.clear

        #データ追加を作る時
        #with open('df.csv','a') as f:
        #    writer = csv.writer(f)
        #    writer.writerow(data)
        #    st.dataframe(writer)
        
        file_id = drive.ListFile({'q': 'title contains "df.csv"'}).GetList()[0]['id']
        a  = drive.CreateFile({'id': file_id,
                            'mimeType':'text/csv'})#ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
        a.GetContentFile('df.csv')
        with open('df.csv','a') as f:
            writer = csv.writer(f)
            writer.writerow(data)
        a.SetContentFile('df.csv')
        a.Upload()
        
       


#編集画面
if genre == '編集':
    st.dataframe(df) 
    #st.text(select)#選択したIDから行のINDEXを取得
    col1, col2 = st.columns(2)
    with col1:
        select = st.selectbox('id', df['id'])
        #st.text(select)#選択したIDから行のINDEXを取得
        #st.text(df['title'][df['id'] == select][])
        #st.text(df['title'][df['id'] == select][2])
        #st.text(df[df['id'] == select].index.to_numpy()[0])
        #main_id2 = df['main_id'][df['id'] == select][0]
        title_2 = df['title'][df['id'] == select].values[0]
        a2 = df['kusa1'][df['id'] == select].values[0]
        b2 = df['kusa2'][df['id'] == select].values[0]
        c2 = df['kusa3'][df['id'] == select].values[0]
        d2 = df['kusa4'][df['id'] == select].values[0]
        e2 = df['kusa5'][df['id'] == select].values[0]
        f2 = df['kuki1'][df['id'] == select].values[0]
        g2 = df['kuki2'][df['id'] == select].values[0]
        h2 = df['kuki3'][df['id'] == select].values[0]
        i2 = df['kuki4'][df['id'] == select].values[0]
        j2 = df['kuki5'][df['id'] == select].values[0]
        k2 = df['spad1'][df['id'] == select].values[0]
        l2 = df['spad2'][df['id'] == select].values[0]
        m2 = df['spad3'][df['id'] == select].values[0]
        n2 = df['spad4'][df['id'] == select].values[0]
        o2 = df['spad5'][df['id'] == select].values[0]
        fx3 = df['field'][df['id'] == select].values[0]
        fb3 = df['close'][df['id'] == select].values[0]
        #tolistでリストが作れる。
        #st.text(df.iloc[0].tolist())
        #main_id = st.text_input('id',value=shoki)
    with col2:
        title3 = st.text_input('タイトル',value=title_2)

    st.markdown('【サンプル１】')
    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        a3 = st.text_input('草丈',value=a2,key=1)#value＝2　デファルト数入れられる
    with col2:
        b3 = st.text_input('',value=b2,key=2)
    with col3:    
        c3 = st.text_input('',value=c2,key=3)
    with col4:
        d3 = st.text_input('',value=d2,key=4)
    with col5:
        e3 = st.text_input('',value=e2,key=5)

    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        f3 = st.text_input('茎数',value=f2,key=6)#value＝2　デファルト数入れられる
    with col2:
        g3 = st.text_input('',value=g2,key=7)
    with col3:    
        h3 = st.text_input('',value=h2,key=8)
    with col4:
        i3 = st.text_input('',value=i2,key=9)
    with col5:
        j3 = st.text_input('',value=j2,key=10)

    col1, col2, col3,col4,col5 = st.columns(5)
    with col1:
        k3 = st.text_input('SPAD',value=k2,key=11)#value＝2　デファルト数入れられる
    with col2:
        l3 = st.text_input('',value=l2,key=12)
    with col3:    
        m3 = st.text_input('',value=m2,key=13)
    with col4:
        n3 = st.text_input('',value=n2,key=14)
    with col5:
        o3 = st.text_input('',value=o2,key=15)

    download_name_a = 'field' + str(select)
    #download_name_a = 'field' + str(df['id'].values[0])
    download_name_b = 'close' + str(select)
    st.text(download_name_a)
    #画像データをクエリをIDかうまくとる方法模索
    for file_ids in drive.ListFile({'q': 'title contains "field"'}).GetList():
    #[0]['id']
        #st.text(file_ids['title'])
        #st.markdown(file_ids['id'])
        if file_ids['title'] == download_name_a:
            file_id = file_ids['id']
        else:
            pass
        #file_ids['id'])
    #st.text(file_id)
    #st.text(file_id)
    f  = drive.CreateFile({'id': file_id})#ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
    f.GetContentFile(download_name_a)
    #st.image(download_name_a)
    #file_id = drive.ListFile({'q': 'title contains "close"'}).GetList()[0]['id']
    for file_ids in drive.ListFile({'q': 'title contains "close"'}).GetList():
    #[0]['id']
        #st.text(file_ids['title'])
        #st.markdown(file_ids['id'])
        if file_ids['title'] == download_name_a:
            file_id = file_ids['id']
        else:
            pass
    f  = drive.CreateFile({'id': file_id})#ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
    f.GetContentFile(download_name_b)
    #
    field = st.file_uploader('全体写真')
    if not field:
        st.image(download_name_a)
    else:
        st.image(field)
    
    close = st.file_uploader('近距離写真')
    if not close:
        st.image(download_name_b)
    else:
        st.image(close)

#テスト既存の写真を消して、今とったものと変える。
    button6 = st.button('field写真変更')
    if button6:
       if field: 
            with open(field.name,'wb') as f:
                #アップロード
                f.write(field.read()) 
                f = drive.CreateFile({'title':download_name_a,#field.name
                            'mimeType':'image/png,image/jpeg',
                            'parents':[{'id':folder_id}]})
                f.SetContentFile(field.name)
                f.Upload()
                #古いデータを消す
                f = drive.CreateFile({'id': fx3})
                f.Trash()
                #IDを取得する。
                file_id_a = drive.ListFile().GetList()
                fx3 = file_id_a['title' == download_name_a]['id']
                st.text(fx3)
       if close:
            with open(close.name,'wb') as f:
                f.write(close.read()) 
                f = drive.CreateFile({'title':download_name_a,#field.name
                            'mimeType':'image/png,image/jpeg',
                                'parents':[{'id':folder_id}]})
                f.SetContentFile(close.name)
                f.Upload()

                #古いデータを消す
                f = drive.CreateFile({'id': fb3})
                f.Trash()

                file_id_a = drive.ListFile().GetList()
                fb3 = file_id_a['title' == download_name_b]['id']


#変更
    henkou = st.button('変更')
    if henkou:
        if field: 
            with open(field.name,'wb') as f:
                #アップロード
                f.write(field.read()) 
                f = drive.CreateFile({'title':download_name_a,#field.name
                            'mimeType':'image/png,image/jpeg',
                            'parents':[{'id':folder_id}]})
                f.SetContentFile(field.name)
                f.Upload()
                #古いデータを消す
                f = drive.CreateFile({'id': fx3})
                f.Trash()
                #IDを取得する。
                file_id_a = drive.ListFile().GetList()
                fx3 = file_id_a['title' == download_name_a]['id']
                st.text(fx3)
        if close:
            with open(close.name,'wb') as f:
                f.write(close.read()) 
                f = drive.CreateFile({'title':download_name_a,#field.name
                            'mimeType':'image/png,image/jpeg',
                                'parents':[{'id':folder_id}]})
                f.SetContentFile(close.name)
                f.Upload()

                #古いデータを消す
                f = drive.CreateFile({'id': fb3})
                f.Trash()

                file_id_a = drive.ListFile().GetList()
                fb3 = file_id_a['title' == download_name_b]['id']



    #変更するデータのリスト
        data_henkou = [select,title3,
        a3,b3,c3,d3,e3,
        f3,g3,h3,i3,j3,
        k3,l3,m3,n3,o3,
        fx3,fb3]
        #st.text(data_henkou)#変更したデータ
        #st.dataframe(df)#変更前のDf
        #st.text(df[df['id'] == select].index.to_numpy()[0])
        k = df[df['id'] == select].index.to_numpy()[0]
        #.index.to_numpy()[0]])
        #st.text(k)
        #st.text(df.iloc[k,:])
        df.iloc[k] = data_henkou
        #st.dataframe(df)
        #st.dataframe(df_d)
    #streamlit 内にあるdf.csvを変更するデータフレームで変更した点をCSVに変換して上書きする
        df.to_csv('df.2csv',index=False)
        file_id = drive.ListFile({'q': 'title contains "df.csv"'}).GetList()[0]['id']
        a  = drive.CreateFile({'id': file_id,
                            'mimeType':'text/csv'})#ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
        a.GetContentFile('df.csv')
        a.SetContentFile('df.2csv')
        a['title'] = 'df.csv'
        a.Upload()
        #a.close
        #with open('df.csv','r')as f:
        #    st.dataframe(f)
#button10 = st.button('消去') 
#if button10:
#    f = drive.CreateFile({'id': '1MRrwVgmmfbRYyKgQwmVmD64_Lf1UHcKB'})
#    f.Trash()
    #st.text(fx3)
    #file = drive_service.files().delete(fileId='1MRrwVgmmfbRYyKgQwmVmD64_Lf1UHcKB').execute()   
#openxlテスト
#import os
#import openpyxl
#from PIL import Image 
#os.chdir('/content/drive/MyDrive')
#ファイルを作成し、データ（数字）を入力

#データ取得


#エクセル化の為
if genre == 'データEXCEL化':
    excel = st.button('excel')
    if excel:
        wb = openpyxl.Workbook()
        #wc = openpyxl.load_workbook('aa.xlsx')
        #ws['A1'] = 10
        wc = wb.active

        
    #画像選択 png形式はRGBA　jpeg形式はRGB
        file_id = drive.ListFile({'q': 'title = "dd.png"'}).GetList()[0]['id']
        f = drive.CreateFile({'id': file_id})#ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
        f.GetContentFile('dd.png')
        
        img_dir = 'dd.png' #width:916,height:685 #width,height = img.size #print(width,height)
        img = Image.open(img_dir).convert('RGB')
    #リサイズ
    #img = Image.open(img_dir)
        img_re = img.resize((300,200))
        img_re.save(img_dir)

        img_to = openpyxl.drawing.image.Image(img_dir)
        wc.add_image(img_to,'B3')
        wb.save('aa.xlsx')

    excel_up = st.button('excel_up')
    if excel_up:
        #フォルダの場所をIDに指定する
        folder_id = '10Ogv7m81vckhXxmRdleo5xouy6lO6O7V'    
        f = drive.CreateFile({'title':'aa.xlsx',#field.name
                            'mimeType':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            'parents':[{'id':folder_id}]})
        f.SetContentFile('aa.xlsx')
        f.Upload()
        f.clear()

#ディレクトリの場所を確認。
#import glob
#st.text(os.getcwd())
#files = glob.glob("/app/google.drive/*")
#for file in files:
#    st.text(file)   

#  Googledriveからデータを取る。
#if  button_download:
    #クエリでlist内の名前で検索、IDを取得。そのIDを使って画像取得
#    file_id = drive.ListFile({'q': 'title = "image2.jpg"'}).GetList()[0]['id']
#    f = drive.CreateFile({'id': file_id})
#    f.GetContentFile(f['title'])
#    st.image(f['title'])
#    f.clear()

#グーグルドライブにあるdf.csvを取得する。
#if  button_download:
#    #クエリでtitle＝ファイル名で検索して、IDを取得する。
#    file_id = drive.ListFile({'q': 'title = "df.csv"'}).GetList()[0]['id']
    #取得したIDでファイル作る。
    #st.text(file_id)
#    f = drive.CreateFile({'id': file_id})
    #ファイルを読み込みして、見えないカレントディレクトリ内に見えないが保存されている。絶対パスで
#    f.GetContentFile('df.csv')
    
    #GOOGLEドライブ内のdf.csvがうまく取り込めないので、DataFrameに変換して取り込む
#    with open('df.csv','r') as f:
#        df = pd.DataFrame(f)
        #t.dataframe(df)
#        list_A = df.iloc[0,0].split(',')
#        list_B= df.iloc[1,0].split(',')
#        data = [list_B]
#        df = pd.DataFrame(data,columns=list_A)
        #df2 = pd.concat([df_i, df_a], axis=0)
#        st.dataframe(df)

#folder_id = '10Ogv7m81vckhXxmRdleo5xouy6lO6O7V'    

#ファイルを一度ドライブの手前のファイルに保存した後にアップロードし、IDでフォルダの場所を指定
#if button_upload:
#if field:
    #st.markdown(f'{field.name}をアップロードしました。')
#    with open(field.name,'wb') as f:
#        f.write(field.read())
    #フォルダの場所をIDに指定する
    #folder_id = '10Ogv7m81vckhXxmRdleo5xouy6lO6O7V'    
#    f = drive.CreateFile({'title':download_name_a,#field.name
#                        'mimeType':'image/png,image/jpeg',
#                        'parents':[{'id':folder_id}]})
#    f.SetContentFile(field.name)
#    f.Upload()

#    file_id_a = drive.ListFile().GetList()
#    fx = file_id_a['title' == download_name_a]['id']
    #st.text(fx)
    #st.write(type(fx))
#    f.clear()
    #st.text(fx)

#if close:
    #st.markdown(f'{close.name}をアップロードしました。')
#    with open(close.name,'wb') as f:
#        f.write(close.read())
#    f = drive.CreateFile({'title':download_name_b,
#                        'mimeType':'image/png,imag/jpeg',
#                        'parents':[{'id':folder_id}]})
#    f.SetContentFile(close.name)
#    f.Upload()
#    file_id_b = drive.ListFile().GetList()
#    fb = file_id_b['title' == download_name_b]['id']
    #st.text(fb)
#    f.clear()
    
#フォルダ作成
#button = st.button('ファルダの作成')
#if button:
#    f_folder = drive.CreateFile({'title':'NEW_Folder',
#                                'mimeType':'application/vnd.google-apps.folder'})
#    f_folder.Upload()    

#テキストをGoogleDriveに保存
#if field:
#    f = drive.CreateFile({'title':'test.txt'})
#    f.SetContentString('test')
#    f.Upload()
#    st.text('アップロード完了')

#データの変換
#    #im = Image.open(field)
#    #im = np.array(im)


#ディレクトを確認。
#st.text(os.getcwd())
#file_list = drive.ListFile().GetList()
#st.text(type(file_list))
# <class 'list'>
#st.text(len(file_list))
# 9
#st.text(type(file_list[0]))
# <class 'pydrive.files.GoogleDriveFile'>
#for f in file_list:
#    st.text(f['title'])
#    st.text(f['id'])

#ディレクトリの場所を確認。
#import glob
#st.text(os.getcwd())
#files = glob.glob("/app/google.drive/*")
#for file in files:
#    st.text(file)   


#if name and field and close:
#    data = [name,im,im2] #ndarray でないとリストに入らないわけでない。
#    st.image(data[1])
#    #tt = Image.fromarray(data[2])
#    st.image(data[2])
#    #Image.open(tt)
#    st.text(data[0])


